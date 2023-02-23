import csv
from openpyxl import load_workbook

wb = load_workbook(filename='Load_header_Import.xlsx')
sheet = wb.active

csv_data=[]
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))

with open('Load_header_Import.csv','w',newline='') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)